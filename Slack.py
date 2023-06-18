import os
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
import pytz
from datetime import datetime
import time
import openpyxl
import requests
import json
import pycountry
import re
from dotenv import load_dotenv

# Set the API token
load_dotenv()
client = WebClient(token=os.getenv('API'))
# Set the API token


# API key for abuse IPdb
api_key = "811f20f61b2f71a30eb3205332868856b5dceef517acb8f243fa310df92936b7725a71a1ec5e6475"

# Set the desired channel ID
channel_id = "Channelid" #careem channel


# Set the reaction emoji
username = input("Please enter your username: ")
if:
 #code
else:
    reaction_emoji = "👋"

# Define the time range in Pakistan time zone
timezone = pytz.timezone('Asia/Karachi')
start_time = timezone.localize(datetime(2023, 5, 23, 0, 0, 0)) # year, month, day, hour, minute, second
end_time = timezone.localize(datetime(2028, 5, 23, 23, 59, 59))

# Load the Excel file
if os.path.exists('message.xlsx'):
    wb = openpyxl.load_workbook('message.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('message.xlsx')
sheet = wb.active
sheet["A1"] = "Timestamp"
sheet["B1"] = "First Message"
sheet["C1"] = "Client"
sheet["D1"] = "Count"
sheet["E1"] = "Reputation"
sheet["F1"] = "Country Name"

# Get the timestamp of the last message that had a reaction added to it
last_message = None
try:
    result = client.conversations_history(channel=channel_id, limit=1)
    if len(result["messages"]) > 0:
        last_message = result["messages"][0]
        if len(last_message.get("reactions", [])) > 0:
            last_timestamp = last_message["ts"]
except SlackApiError as e:
    print("Error getting last message timestamp: {}".format(e))

# Initialize the list of processed message timestamps
processed_messages = []

while True:
    try:
        # Call the conversations.history method using the WebClient with the oldest and latest timestamp parameters
        oldest_timestamp = last_message["ts"] if last_message else start_time.timestamp()
        latest_result = client.conversations_history(channel=channel_id, oldest=oldest_timestamp, latest=end_time.timestamp())

        # Iterate over each new message and add the reaction if it falls within the time range
        for message in latest_result["messages"]:
            message_time = datetime.fromtimestamp(float(message["ts"]), tz=pytz.utc).astimezone(timezone)
            time_diff = datetime.now(timezone) - message_time
            if time_diff.total_seconds() <= 10 and message["ts"] not in processed_messages:
                if len(message.get("reactions", [])) == 0:
                    # Fetch the IP and first line of the message
                    client_ip = None
                    first_line = None
                    countryname= None
                    count = 1
                    for i, line in enumerate(message["text"].splitlines()):
                       # print (line)
                        if "client:" in line:
                            print (line)
                            if i + 1 < len(message["text"].splitlines()):
                                next_line = message["text"].splitlines()[i + 1]
                                ip_address = re.search(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', line)
                                if ip_address:
                                    client_ip = ip_address.group()
                                    print(client_ip)
                                else:
                                    ip_address = re.search(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', next_line)
                                    if ip_address:
                                        client_ip = ip_address.group()
                                        print(client_ip)
                               
                           
                        
                            url = f"https://api.abuseipdb.com/api/v2/check?ipAddress={client_ip}"
                            headers = {"Key": api_key, "Accept": "application/json"}
                            response = requests.get(url, headers=headers)
                            if response.status_code == 200:
                              result = json.loads(response.text)["data"]
                            if result["abuseConfidenceScore"]:
                              print(f"IP address: {client_ip}\nReputation: {result['abuseConfidenceScore']}/100")
                            else:
                             print(f"IP address: {client_ip}\nReputation: Not available")
                            if result["countryCode"]:
                                country_code = result['countryCode']
                                print(f"Country Code: {country_code}")
                                try:
                                    country_name = pycountry.countries.get(alpha_2=country_code).name
                                    countryname = country_name
                                    print(f"Country Name: {country_name}")
                                except Exception as e:
                                     print(f"Error: Unable to retrieve country information from AbuseIPDB. {e}")
                            else:
                                print("Country: Not available")
                        elif first_line is None:
                            first_line = line.strip()


                    # Add the reaction and print a success message
                    client.reactions_add(channel=channel_id, timestamp=message["ts"], name=reaction_emoji)
                    print("Reaction added successfully for message {} in channel {}!".format(message["ts"], channel_id))

                    # checking for the duplication of ip addresses

                    # get the column headers
                    headers = [cell.value for cell in sheet[1]]
                    # get the last row index
                    last_row_index = sheet.max_row

                    # get the IP address from the new data
                    new_first_line= first_line
                    new_ip = client_ip
                    # check if the IP address already exists in the worksheet
                    ip_index = None
                    for i in range(2, last_row_index+1):
                       if sheet.cell(row=i, column=headers.index('Client')+1).value == new_ip and sheet.cell(row=i, column=headers.index('First Message')+1).value == new_first_line:
                         ip_index = i
                         break
                    # update the count value if the IP address exists, or add a new row otherwise
                    if ip_index:
                        count = sheet.cell(row=ip_index, column=headers.index('Count')+1).value + 1
                        old_row_index = ip_index
                        sheet.cell(row=ip_index, column=headers.index('Count')+1).value = count
                        sheet.cell(row=ip_index, column=headers.index('Timestamp')+1).value = message_time.strftime('%Y-%m-%d %H:%M:%S %Z%z')
                        row = sheet[ip_index]
                        sheet.delete_rows(old_row_index)
                        sheet.append(row)
                       # sheet.move_range(f"A{ip_index}:F{ip_index}", rows=last_row_index-1)
                    # Write first line and client IP to Excel file                   
                    else:
                        sheet.append([message_time.strftime('%Y-%m-%d %H:%M:%S %Z%z'), first_line, client_ip, count, result['abuseConfidenceScore'], countryname])

                    wb.save('message.xlsx')
                    
                    
                processed_messages.append(message["ts"])

        # Update the last message
        last_message = latest_result["messages"][-1] if len(latest_result["messages"]) > 0 else last_message

    except SlackApiError as e:
        print("Error adding reactions: {}".format(e))

    # Wait for 180 seconds before checking for new messages again
    time.sleep(10)

